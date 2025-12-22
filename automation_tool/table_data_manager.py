#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
表格数据管理器模块
"""
import openpyxl
from typing import List, Dict, Optional
from tkinter import messagebox


class TableDataManager:
    """表格数据管理器，负责读取表格文件并管理变量"""
    
    def __init__(self):
        """初始化表格数据管理器"""
        self.file_path = ""
        self.data = []
        self.fields = []
        self.current_record = None
        self.current_index = -1
    
    def load_table(self, file_path: str) -> bool:
        """
        加载表格文件
        :param file_path: 表格文件路径
        :return: 加载成功返回True，失败返回False
        """
        try:
            self.file_path = file_path
            
            # 根据文件扩展名选择不同的读取方式
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                return self._load_excel(file_path)
            else:
                messagebox.showerror("错误", "不支持的文件格式，仅支持Excel文件(.xlsx, .xls)")
                return False
        except Exception as e:
            print(f"加载表格文件失败: {e}")
            messagebox.showerror("错误", f"加载表格文件失败: {str(e)}")
            return False
    
    def _load_excel(self, file_path: str) -> bool:
        """
        加载Excel文件
        :param file_path: Excel文件路径
        :return: 加载成功返回True，失败返回False
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # 读取表头作为字段名
            self.fields = []
            for col in range(1, sheet.max_column + 1):
                field_name = sheet.cell(row=1, column=col).value
                if field_name:
                    self.fields.append(str(field_name).strip())
                else:
                    self.fields.append(f"字段{col}")
            
            # 读取数据行
            self.data = []
            for row in range(2, sheet.max_row + 1):
                row_data = {}
                has_data = False
                for col in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=row, column=col).value
                    field_name = self.fields[col - 1]
                    row_data[field_name] = str(value) if value is not None else ""
                    if value is not None:
                        has_data = True
                if has_data:
                    self.data.append(row_data)
            
            workbook.close()
            self.current_index = -1
            self.current_record = None
            return True
        except Exception as e:
            print(f"读取Excel文件失败: {e}")
            raise
    
    def get_next_record(self) -> Optional[Dict[str, str]]:
        """
        获取下一条记录
        :return: 下一条记录字典，没有更多记录返回None
        """
        self.current_index += 1
        if self.current_index < len(self.data):
            self.current_record = self.data[self.current_index]
            return self.current_record
        return None
    
    def reset(self) -> None:
        """
        重置记录指针
        """
        self.current_index = -1
        self.current_record = None
    
    def get_fields(self) -> List[str]:
        """
        获取所有字段名
        :return: 字段名列表
        """
        return self.fields.copy()
    
    def get_total_records(self) -> int:
        """
        获取总记录数
        :return: 总记录数
        """
        return len(self.data)
