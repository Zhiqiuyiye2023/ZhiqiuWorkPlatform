"""
文件批处理处理器
作者: 知秋一叶
版本号: 0.0.5
"""

import os
import shutil
import pandas as pd
from typing import List, Tuple


class FileBatchProcessor:
    """文件批处理处理器"""
    
    def __init__(self):
        """初始化处理器"""
        pass
    
    def load_copy_from_excel(self, excel_path: str, source_col: str, target_col: str) -> List[Tuple[str, str]]:
        """
        从Excel加载复制/移动数据
        
        Args:
            excel_path: Excel文件路径
            source_col: 源路径列名
            target_col: 目标路径列名
            
        Returns:
            包含(源路径, 目标路径)元组的列表
        """
        df = pd.read_excel(excel_path)
        copy_data = []
        
        # 如果只选择了目标列，则源路径使用Excel文件所在目录
        if source_col and not target_col:
            source_base = os.path.dirname(excel_path)
            for _, row in df.iterrows():
                target_path = str(row[target_col])
                copy_data.append((source_base, target_path))
        elif target_col and not source_col:
            # 如果只选择了目标列，则源路径使用Excel文件所在目录
            source_base = os.path.dirname(excel_path)
            for _, row in df.iterrows():
                target_path = str(row[target_col])
                copy_data.append((source_base, target_path))
        else:
            # 同时选择了源和目标列
            for _, row in df.iterrows():
                source_path = str(row[source_col])
                target_path = str(row[target_col])
                copy_data.append((source_path, target_path))
                
        return copy_data
    
    def load_copy_from_text(self, source_text: str, target_text: str) -> List[Tuple[str, str]]:
        """
        从文本加载复制/移动数据
        
        Args:
            source_text: 源路径文本（每行一个）
            target_text: 目标路径文本（每行一个）
            
        Returns:
            包含(源路径, 目标路径)元组的列表
        """
        source_lines = source_text.strip().split('\n')
        target_lines = target_text.strip().split('\n')
        
        # 清理空行
        source_lines = [line.strip() for line in source_lines if line.strip()]
        target_lines = [line.strip() for line in target_lines if line.strip()]
        
        copy_data = []
        for source_path, target_path in zip(source_lines, target_lines):
            copy_data.append((source_path, target_path))
            
        return copy_data
    
    def execute_copy(self, copy_data: List[Tuple[str, str]]) -> Tuple[int, int, List[Tuple[str, str, str, str]]]:
        """
        执行批量复制操作
        
        Args:
            copy_data: 包含(源路径, 目标路径)元组的列表
            
        Returns:
            (成功数量, 失败数量, 处理过的项目列表)
            处理过的项目列表包含(源路径, 目标路径, 状态, 结果信息)元组
            状态: '成功' 或 '失败'
        """
        success_count = 0
        fail_count = 0
        processed_items = []
        
        for source_path, target_path in copy_data:
            try:
                # 确保目标目录存在
                target_dir = os.path.dirname(target_path)
                if target_dir and not os.path.exists(target_dir):
                    os.makedirs(target_dir)
                
                # 执行复制
                if os.path.isfile(source_path):
                    shutil.copy2(source_path, target_path)
                elif os.path.isdir(source_path):
                    if os.path.exists(target_path):
                        shutil.rmtree(target_path)
                    shutil.copytree(source_path, target_path)
                else:
                    raise FileNotFoundError(f"源路径不存在: {source_path}")
                    
                success_count += 1
                processed_items.append((source_path, target_path, '成功', ''))
            except Exception as e:
                fail_count += 1
                error_msg = f"复制失败 {source_path} -> {target_path}: {str(e)}"
                print(error_msg)
                processed_items.append((source_path, target_path, '失败', str(e)))
                
        return success_count, fail_count, processed_items
    
    def execute_move(self, copy_data: List[Tuple[str, str]]) -> Tuple[int, int, List[Tuple[str, str, str, str]]]:
        """
        执行批量移动操作
        
        Args:
            copy_data: 包含(源路径, 目标路径)元组的列表
            
        Returns:
            (成功数量, 失败数量, 处理过的项目列表)
            处理过的项目列表包含(源路径, 目标路径, 状态, 结果信息)元组
            状态: '成功' 或 '失败'
        """
        success_count = 0
        fail_count = 0
        processed_items = []
        
        for source_path, target_path in copy_data:
            try:
                # 确保目标目录存在
                target_dir = os.path.dirname(target_path)
                if target_dir and not os.path.exists(target_dir):
                    os.makedirs(target_dir)
                
                # 执行移动
                shutil.move(source_path, target_path)
                success_count += 1
                processed_items.append((source_path, target_path, '成功', ''))
            except Exception as e:
                fail_count += 1
                error_msg = f"移动失败 {source_path} -> {target_path}: {str(e)}"
                print(error_msg)
                processed_items.append((source_path, target_path, '失败', str(e)))
                
        return success_count, fail_count, processed_items
    
    def batch_create_folders(self, folder_paths: List[str]) -> Tuple[int, int, List[Tuple[str, str, str]]]:
        """
        批量创建文件夹
        
        Args:
            folder_paths: 文件夹路径列表
            
        Returns:
            (成功数量, 失败数量, 处理过的项目列表)
            处理过的项目列表包含(文件夹路径, 状态, 结果信息)元组
            状态: '成功' 或 '失败'
        """
        success_count = 0
        fail_count = 0
        created_folders = []
        
        for folder_path in folder_paths:
            try:
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                    success_count += 1
                    created_folders.append((folder_path, '成功', ''))
                else:
                    # 文件夹已存在
                    success_count += 1
                    created_folders.append((folder_path, '成功', '文件夹已存在'))
            except Exception as e:
                fail_count += 1
                error_msg = f"创建文件夹失败 {folder_path}: {str(e)}"
                print(error_msg)
                created_folders.append((folder_path, '失败', str(e)))
                
        return success_count, fail_count, created_folders
    
    def load_rename_from_excel(self, excel_path: str, old_col: str, new_col: str) -> List[Tuple[str, str]]:
        """
        从Excel加载重命名数据
        
        Args:
            excel_path: Excel文件路径
            old_col: 原名称列名
            new_col: 新名称列名
            
        Returns:
            包含(原名称, 新名称)元组的列表
        """
        df = pd.read_excel(excel_path)
        rename_data = []
        
        for _, row in df.iterrows():
            old_name = str(row[old_col])
            new_name = str(row[new_col])
            rename_data.append((old_name, new_name))
            
        return rename_data
    
    def load_rename_from_text(self, text: str) -> List[Tuple[str, str]]:
        """
        从文本加载重命名数据
        
        Args:
            text: 文本内容（每行格式：原名称<TAB>新名称）
            
        Returns:
            包含(原名称, 新名称)元组的列表
        """
        lines = text.strip().split('\n')
        rename_data = []
        
        for line in lines:
            parts = line.strip().split('\t')
            if len(parts) >= 2:
                old_name = parts[0]
                new_name = parts[1]
                rename_data.append((old_name, new_name))
                
        return rename_data
    
    def load_data_from_excel(self, excel_path: str, id_col: str, content_col: str) -> str:
        """
        从Excel加载数据处理数据
        
        Args:
            excel_path: Excel文件路径
            id_col: 序号列名
            content_col: 关联内容列名
            
        Returns:
            格式化文本内容
        """
        df = pd.read_excel(excel_path)
        lines = []
        
        for _, row in df.iterrows():
            id_value = str(row[id_col])
            content_value = str(row[content_col])
            lines.append(f"{id_value}\t{content_value}")
            
        return '\n'.join(lines)
    
    def process_data(self, text_content: str) -> str:
        """
        处理数据（将相同序号的内容合并）
        
        Args:
            text_content: 输入文本内容（每行格式：序号<TAB>内容）
            
        Returns:
            处理后的文本内容
        """
        lines = text_content.strip().split('\n')
        data_dict = {}
        
        # 按序号分组
        for line in lines:
            parts = line.strip().split('\t')
            if len(parts) >= 2:
                id_value = parts[0]
                content_value = parts[1]
                if id_value in data_dict:
                    data_dict[id_value].append(content_value)
                else:
                    data_dict[id_value] = [content_value]
        
        # 生成结果
        result_lines = []
        for id_value, content_list in data_dict.items():
            merged_content = ','.join(content_list)
            result_lines.append(f"{id_value}\t{merged_content}")
            
        return '\n'.join(result_lines)
    
    def batch_create_folders_old(self, target_path: str, folder_names: str) -> List[str]:
        """
        批量创建文件夹（旧方法）
        
        Args:
            target_path: 目标路径
            folder_names: 文件夹名称（每行一个）
            
        Returns:
            创建的文件夹路径列表
        """
        names = folder_names.strip().split('\n')
        names = [name.strip() for name in names if name.strip()]
        created_folders = []
        
        for name in names:
            folder_path = os.path.join(target_path, name)
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)
                created_folders.append(folder_path)
                
        return created_folders
    
    def move_file_method(self, source_path: str, target_path: str, file_names: str) -> List[str]:
        """
        移动文件方法
        
        Args:
            source_path: 源路径
            target_path: 目标路径
            file_names: 文件名称（每行一个）
            
        Returns:
            移动的文件路径列表
        """
        names = file_names.strip().split('\n')
        names = [name.strip() for name in names if name.strip()]
        moved_files = []
        
        for name in names:
            source_file = os.path.join(source_path, name)
            target_file = os.path.join(target_path, name)
            
            if os.path.exists(source_file):
                # 确保目标目录存在
                target_dir = os.path.dirname(target_file)
                if target_dir and not os.path.exists(target_dir):
                    os.makedirs(target_dir)
                
                shutil.move(source_file, target_file)
                moved_files.append(target_file)
                
        return moved_files
    
    def execute_delete(self, delete_data: List[str]) -> Tuple[int, int, List[Tuple[str, str, str, str]]]:
        """
        执行批量删除操作
        
        Args:
            delete_data: 要删除的路径列表
            
        Returns:
            (成功数量, 失败数量, 处理过的项目列表)
            处理过的项目列表包含(源路径, 目标路径, 状态, 结果信息)元组
            状态: '成功' 或 '失败'
        """
        success_count = 0
        fail_count = 0
        processed_items = []
        
        for path in delete_data:
            try:
                # 执行删除
                if os.path.isfile(path):
                    os.remove(path)
                elif os.path.isdir(path):
                    shutil.rmtree(path)
                else:
                    raise FileNotFoundError(f"路径不存在: {path}")
                    
                success_count += 1
                processed_items.append((path, "", '成功', ''))
            except Exception as e:
                fail_count += 1
                error_msg = f"删除失败: {str(e)}"
                print(error_msg)
                processed_items.append((path, "", '失败', str(e)))
                
        return success_count, fail_count, processed_items
    
    def compare_full_tables(self, excel1_path: str, excel2_path: str) -> Tuple[List, List, List]:
        """
        比对两个表格的完整内容
        
        Args:
            excel1_path: 第一个Excel文件路径
            excel2_path: 第二个Excel文件路径
            
        Returns:
            (比对结果列表, 差异单元格列表, 列名列表)
        """
        # 读取两个Excel文件
        df1 = pd.read_excel(excel1_path)
        df2 = pd.read_excel(excel2_path)
        
        # 比较列名
        columns1 = set(df1.columns)
        columns2 = set(df2.columns)
        
        # 计算列差异
        only_in1_cols = columns1 - columns2
        only_in2_cols = columns2 - columns1
        
        index = 1
        results = []
        different_cells = []  # 存储差异单元格的位置 (行号, 列名)
        
        # 显示列差异
        for col in only_in1_cols:
            results.append((index, col, "", "仅第一个表格有此列"))
            index += 1
        
        for col in only_in2_cols:
            results.append((index, "", col, "仅第二个表格有此列"))
            index += 1
        
        # 比较共同列的数据
        common_cols = columns1 & columns2
        
        if common_cols:
            # 确定最小行数
            min_rows = min(len(df1), len(df2))
            
            # 比较每一行每一列的数据
            for i in range(min_rows):
                for col in common_cols:
                    val1 = df1.iloc[i][col]
                    val2 = df2.iloc[i][col]
                    
                    # 处理NaN值
                    if pd.isna(val1) and pd.isna(val2):
                        continue
                    if pd.isna(val1) or pd.isna(val2):
                        # 一个有值一个无值，标记为差异
                        different_cells.append((i, col))
                        # 显示差异
                        results.append((index, f"行{i+1} {col}: {val1}", f"行{i+1} {col}: {val2}", "内容不一致"))
                        index += 1
                    elif str(val1) != str(val2):
                        # 内容不一致，标记为差异
                        different_cells.append((i, col))
                        # 显示差异
                        results.append((index, f"行{i+1} {col}: {val1}", f"行{i+1} {col}: {val2}", "内容不一致"))
                        index += 1
            
            # 处理行数差异
            if len(df1) != len(df2):
                # 显示行数差异
                results.append((index, f"总行数: {len(df1)}", f"总行数: {len(df2)}", "行数不同"))
                index += 1
        
        return results, different_cells, list(df1.columns)
    
    def mark_differences_in_excel(self, excel_path: str, different_cells: List, columns_list: List, 
                                  is_cell_bg: bool = True, mark_color: str = "红色") -> Tuple[str, int]:
        """
        在Excel文件中标红差异并保存
        
        Args:
            excel_path: Excel文件路径
            different_cells: 差异单元格列表
            columns_list: 列名列表
            is_cell_bg: 是否标红单元格背景（True）或文字颜色（False）
            mark_color: 标红颜色
            
        Returns:
            (输出文件路径, 标红数量) 或 None
        """
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        try:
            # 颜色映射字典
            color_map = {
                "红色": "FFFF0000",
                "蓝色": "FF0000FF",
                "黄色": "FFFFFF00",
                "绿色": "FF00FF00",
                "橙色": "FFFFA500"
            }
            
            # 获取颜色十六进制值
            color_hex = color_map.get(mark_color, "FFFF0000")
            
            # 加载Excel文件
            workbook = load_workbook(excel_path)
            sheet = workbook.active
            
            # 遍历差异单元格，根据选择的方式标红
            marked_count = 0
            
            for row_idx, col_name in different_cells:
                try:
                    # 获取列索引
                    if col_name in columns_list:
                        col_idx = columns_list.index(col_name) + 1  # 列索引从1开始
                        
                        # 计算Excel行号：pandas的行索引+2（因为Excel从1开始，且有1行表头）
                        excel_row = row_idx + 2
                        
                        # 检查行是否存在
                        if excel_row <= sheet.max_row and col_idx <= sheet.max_column:
                            # 获取单元格
                            cell = sheet.cell(row=excel_row, column=col_idx)
                            
                            # 根据选择的标红方式应用不同的样式
                            if is_cell_bg:
                                # 标红单元格背景
                                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                                cell.fill = fill
                            else:
                                # 标红文字颜色
                                font = Font(color=color_hex)
                                cell.font = font
                            
                            marked_count += 1
                except Exception as cell_e:
                    print(f"标红单元格时出错: {cell_e}")
            
            # 保存标红文件
            output_path = excel_path.replace('.xlsx', '_标红.xlsx')
            workbook.save(output_path)
            
            return output_path, marked_count
        except Exception as e:
            print(f"标红文件时出错: {str(e)}")
            return None