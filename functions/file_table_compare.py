"""
表格比对功能
作者: 知秋一叶
版本号: 0.0.6
"""

from PyQt6.QtWidgets import QHBoxLayout, QVBoxLayout, QLabel, QFileDialog, QTableWidgetItem, QHeaderView, QMessageBox, QRadioButton, QButtonGroup
from qfluentwidgets import LineEdit, PushButton, TableWidget, ComboBox
from qfluentwidgets import FluentIcon as FIF
from .file_processor_base import BaseFileProcessorFunction
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from .file_processor_base import BaseFileProcessorFunction


class FileTableCompareFunction(BaseFileProcessorFunction):
    """表格比对功能"""
    
    def __init__(self, parent=None):
        description = (
            "📢 <b>功能说明：</b><br>"
            "比对两个Excel表格内容<br>"
            "支持查找差异项并高亮显示"
        )
        super().__init__("表格比对", description, parent)
        self._initUI()
    
    def _initUI(self):
        """初始化界面"""
        # 第一个Excel文件选择区域
        file1_layout = QHBoxLayout()
        file1_layout.addWidget(QLabel("第一个Excel文件:"))
        
        self.file1_edit = LineEdit(self)
        self.file1_edit.setPlaceholderText("请选择第一个Excel文件")
        
        self.file1_browse_button = PushButton("选择文件", self, FIF.DOCUMENT)
        self.file1_browse_button.clicked.connect(lambda: self.browse_file(self.file1_edit, "Excel Files (*.xlsx *.xls)"))
        
        file1_layout.addWidget(self.file1_edit)
        file1_layout.addWidget(self.file1_browse_button)
        self.contentLayout.addLayout(file1_layout)
        
        # 第二个Excel文件选择区域
        file2_layout = QHBoxLayout()
        file2_layout.addWidget(QLabel("第二个Excel文件:"))
        
        self.file2_edit = LineEdit(self)
        self.file2_edit.setPlaceholderText("请选择第二个Excel文件")
        
        self.file2_browse_button = PushButton("选择文件", self, FIF.DOCUMENT)
        self.file2_browse_button.clicked.connect(lambda: self.browse_file(self.file2_edit, "Excel Files (*.xlsx *.xls)"))
        
        file2_layout.addWidget(self.file2_edit)
        file2_layout.addWidget(self.file2_browse_button)
        self.contentLayout.addLayout(file2_layout)
        
        # 标红设置区域
        mark_settings_layout = QHBoxLayout()
        mark_settings_layout.addWidget(QLabel("标红方式:"))
        
        # 标红方式选择（文字颜色/单元格背景）
        self.mark_type_group = QButtonGroup(self)
        self.text_color_radio = QRadioButton("标红文字")
        self.cell_bg_radio = QRadioButton("标红单元格")
        self.cell_bg_radio.setChecked(True)  # 默认标红单元格
        
        self.mark_type_group.addButton(self.text_color_radio)
        self.mark_type_group.addButton(self.cell_bg_radio)
        
        mark_settings_layout.addWidget(self.text_color_radio)
        mark_settings_layout.addWidget(self.cell_bg_radio)
        
        # 颜色选择
        mark_settings_layout.addWidget(QLabel("标红颜色:"))
        self.color_combo = ComboBox(self)
        self.color_combo.addItems(["红色", "蓝色", "黄色", "绿色", "橙色"])
        self.color_combo.setCurrentText("红色")  # 默认红色
        mark_settings_layout.addWidget(self.color_combo)
        
        mark_settings_layout.addStretch()
        self.contentLayout.addLayout(mark_settings_layout)
        
        # 比对结果显示区域
        self.result_table = TableWidget(self)
        self.result_table.setColumnCount(4)
        self.result_table.setHorizontalHeaderLabels(["序号", "第一个表格内容", "第二个表格内容", "差异类型"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setFixedHeight(300)
        self.result_table.setBorderVisible(True)
        self.contentLayout.addWidget(self.result_table)
        
        # 按钮区域
        button_layout = QHBoxLayout()
        self.compare_button = PushButton("比对表格", self, FIF.SEARCH)
        self.export_button = PushButton("导出结果", self, FIF.SAVE)
        
        self.compare_button.clicked.connect(self.compare_tables)
        self.export_button.clicked.connect(self.export_results)
        
        button_layout.addWidget(self.compare_button)
        button_layout.addWidget(self.export_button)
        self.contentLayout.addLayout(button_layout)
    
    def compare_tables(self):
        """比对表格"""
        file1_path = self.file1_edit.text()
        file2_path = self.file2_edit.text()
        
        if not file1_path:
            self.show_warning("警告", "请选择第一个Excel文件")
            return
            
        if not file2_path:
            self.show_warning("警告", "请选择第二个Excel文件")
            return
            
        try:
            self.showProgress("正在比对表格...")
            # 读取两个Excel文件
            df1 = pd.read_excel(file1_path)
            df2 = pd.read_excel(file2_path)
            
            # 清空结果表格
            self.result_table.setRowCount(0)
            
            # 比较列名
            columns1 = set(df1.columns)
            columns2 = set(df2.columns)
            
            # 计算列差异
            only_in1_cols = columns1 - columns2
            only_in2_cols = columns2 - columns1
            
            index = 1
            results = []
            different_cells = []  # 存储差异单元格的位置 (行号, 列名)
            
            # 显示列差异
            for col in only_in1_cols:
                results.append((index, col, "", "仅第一个表格有此列"))
                index += 1
            
            for col in only_in2_cols:
                results.append((index, "", col, "仅第二个表格有此列"))
                index += 1
            
            # 比较共同列的数据
            common_cols = columns1 & columns2
            
            if common_cols:
                # 确定最小行数
                min_rows = min(len(df1), len(df2))
                
                # 比较每一行每一列的数据
                for i in range(min_rows):
                    for col in common_cols:
                        val1 = df1.iloc[i][col]
                        val2 = df2.iloc[i][col]
                        
                        # 处理NaN值
                        if pd.isna(val1) and pd.isna(val2):
                            continue
                        if pd.isna(val1) or pd.isna(val2):
                            # 一个有值一个无值，标记为差异
                            different_cells.append((i, col))
                            # 显示差异
                            results.append((index, f"行{i+1} {col}: {val1}", f"行{i+1} {col}: {val2}", "内容不一致"))
                            index += 1
                        elif str(val1) != str(val2):
                            # 内容不一致，标记为差异
                            different_cells.append((i, col))
                            # 显示差异
                            results.append((index, f"行{i+1} {col}: {val1}", f"行{i+1} {col}: {val2}", "内容不一致"))
                            index += 1
                
                # 处理行数差异
                if len(df1) != len(df2):
                    # 显示行数差异
                    results.append((index, f"总行数: {len(df1)}", f"总行数: {len(df2)}", "行数不同"))
                    index += 1
            
            # 显示结果
            self.display_results(results)
            
            # 在第一个表格中标红差异并保存
            mark_success = False
            output_path = ""
            marked_count = 0
            
            if different_cells:
                mark_success, output_path, marked_count = self.mark_differences(file1_path, df1, different_cells)
            
            # 只显示一个整合的成功消息
            if mark_success:
                self.showSuccess(f"比对完成，发现 {len(results)} 处差异\n差异已标红并保存到: {output_path}\n成功标红: {marked_count} 个单元格")
            else:
                self.showSuccess(f"比对完成，发现 {len(results)} 处差异")
        except Exception as e:
            self.showError(f"比对时出错: {str(e)}")
        finally:
            self.hideProgress()
    
    def display_results(self, results):
        """显示比对结果"""
        self.result_table.setRowCount(len(results))
        for i, (index, content1, content2, diff_type) in enumerate(results):
            # 序号列
            index_item = QTableWidgetItem(str(index))
            self.result_table.setItem(i, 0, index_item)
            
            # 第一个表格内容列
            content1_item = QTableWidgetItem(str(content1))
            self.result_table.setItem(i, 1, content1_item)
            
            # 第二个表格内容列
            content2_item = QTableWidgetItem(str(content2))
            self.result_table.setItem(i, 2, content2_item)
            
            # 差异类型列
            diff_type_item = QTableWidgetItem(diff_type)
            self.result_table.setItem(i, 3, diff_type_item)
    
    def mark_differences(self, file_path, df, different_cells):
        """在Excel文件中标红差异并保存"""
        try:
            # 颜色映射字典
            color_map = {
                "红色": "FFFF0000",
                "蓝色": "FF0000FF",
                "黄色": "FFFFFF00",
                "绿色": "FF00FF00",
                "橙色": "FFFFA500"
            }
            
            # 获取用户选择的标红方式和颜色
            is_cell_bg = self.cell_bg_radio.isChecked()
            selected_color = self.color_combo.currentText()
            color_hex = color_map.get(selected_color, "FFFF0000")
            
            # 加载Excel文件
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            # 遍历差异单元格，根据选择的方式标红
            marked_count = 0
            
            for row_idx, col_name in different_cells:
                try:
                    # 获取列索引，避免重新匹配列名
                    if col_name in df.columns:
                        col_idx = df.columns.get_loc(col_name) + 1  # 列索引从1开始
                        
                        # 计算Excel行号：pandas的行索引+2（因为Excel从1开始，且有1行表头）
                        excel_row = row_idx + 2
                        
                        # 检查行是否存在
                        if excel_row <= sheet.max_row and col_idx <= sheet.max_column:
                            # 获取单元格
                            cell = sheet.cell(row=excel_row, column=col_idx)
                            
                            # 根据选择的标红方式应用不同的样式
                            if is_cell_bg:
                                # 标红单元格背景
                                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                                cell.fill = fill
                            else:
                                # 标红文字颜色
                                font = Font(color=color_hex)
                                cell.font = font
                            
                            marked_count += 1
                except Exception as cell_e:
                    print(f"标红单元格时出错: {cell_e}")
            
            # 保存标红文件
            output_path = file_path.replace('.xlsx', '_标红.xlsx')
            workbook.save(output_path)
            
            return True, output_path, marked_count
        except Exception as e:
            print(f"标红文件时出错: {str(e)}")
            return False, None, 0
    
    def export_results(self):
        """导出比对结果"""
        if self.result_table.rowCount() == 0:
            self.show_warning("警告", "没有比对结果可导出")
            return
            
        try:
            file_path, _ = QFileDialog.getSaveFileName(self, "保存结果", "", "Excel Files (*.xlsx)")
            if file_path:
                # 创建DataFrame
                data = []
                for i in range(self.result_table.rowCount()):
                    row = []
                    for j in range(self.result_table.columnCount()):
                        item = self.result_table.item(i, j)
                        row.append(item.text() if item else "")
                    data.append(row)
                
                df = pd.DataFrame(data, columns=["序号", "第一个表格内容", "第二个表格内容", "差异类型"])
                df.to_excel(file_path, index=False)
                
                self.show_success("成功", "结果已导出")
        except Exception as e:
            self.show_error("错误", f"导出结果时出错: {str(e)}")
