# coding:utf-8
"""
表格合并功能
用于合并多个表格文件，支持拖入文件和添加目录
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PyQt6.QtWidgets import (
    QHBoxLayout, QVBoxLayout, QLabel, QFileDialog, 
    QTableWidgetItem, QMessageBox, QGroupBox, QCheckBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QDragEnterEvent, QDropEvent
from qfluentwidgets import (
    PushButton, ComboBox, LineEdit, 
    InfoBar, InfoBarPosition, PrimaryPushButton, 
    FluentIcon, TableWidget, BodyLabel
)
from .file_processor_base import BaseFileProcessorFunction


def merge_tables(file_list, output_path, merge_type='vertical', preserve_format=True, header_rows=0):
    """
    合并多个表格文件
    
    Args:
        file_list (list): 表格文件路径列表
        output_path (str): 输出文件路径
        merge_type (str): 合并方式，'vertical'（垂直合并）或 'horizontal'（水平合并）
        preserve_format (bool): 是否保留原表格格式
        header_rows (int): 表头行数，0表示保留所有表头
    """
    if not file_list:
        return False, "没有要合并的表格文件"
    
    try:
        if preserve_format and all(f.lower().endswith('.xlsx') for f in file_list):
            # 如果所有文件都是xlsx格式且需要保留格式，使用openpyxl处理
            return merge_tables_with_format(file_list, output_path, merge_type, header_rows)
        else:
            # 否则使用pandas处理
            return merge_tables_pandas(file_list, output_path, merge_type, header_rows)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return False, f"合并过程中发生错误：{str(e)}"


def merge_tables_pandas(file_list, output_path, merge_type='vertical', header_rows=0):
    """
    使用pandas合并表格
    
    Args:
        file_list (list): 表格文件路径列表
        output_path (str): 输出文件路径
        merge_type (str): 合并方式，'vertical'（垂直合并）或 'horizontal'（水平合并）
        header_rows (int): 表头行数，0表示保留所有表头（仅垂直合并有效）
    """
    # 读取所有表格
    dataframes = []
    for i, file_path in enumerate(file_list):
        if file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xls'):
            df = pd.read_excel(file_path)
        elif file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            return False, f"不支持的文件格式：{os.path.basename(file_path)}"
        
        # 只有垂直合并且设置了表头行数时，才处理表头
        if merge_type == 'vertical' and i > 0 and header_rows > 0:
            dataframes.append(df)
        else:
            dataframes.append(df)
    
    # 执行合并
    if merge_type == 'vertical':
        # 垂直合并（追加行）
        merged_df = pd.concat(dataframes, ignore_index=True)
    else:
        # 水平合并（追加列）
        # 简单拼接，忽略表头行数设置
        merged_df = pd.concat(dataframes, axis=1)
    
    # 保存结果
    if output_path.lower().endswith('.xlsx'):
        merged_df.to_excel(output_path, index=False)
    elif output_path.lower().endswith('.csv'):
        merged_df.to_csv(output_path, index=False, encoding='utf-8-sig')
    else:
        # 默认保存为Excel
        output_path += '.xlsx'
        merged_df.to_excel(output_path, index=False)
    
    return True, f"合并完成！输出文件：{output_path}"


def merge_tables_with_format(file_list, output_path, merge_type='vertical', header_rows=0):
    """
    使用openpyxl合并表格并保留格式
    
    Args:
        file_list (list): 表格文件路径列表
        output_path (str): 输出文件路径
        merge_type (str): 合并方式，'vertical'（垂直合并）或 'horizontal'（水平合并）
        header_rows (int): 表头行数，0表示保留所有表头（仅垂直合并有效）
    """
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from copy import copy
    
    # 创建新的工作簿
    wb_out = Workbook()
    ws_out = wb_out.active
    
    if merge_type == 'vertical':
        # 垂直合并（追加行）
        row_offset = 0
        for i, file_path in enumerate(file_list):
            wb_in = load_workbook(file_path)
            ws_in = wb_in.active
            
            # 读取源工作表的格式信息
            formats = read_worksheet_formats(ws_in)
            
            # 确定实际有数据的行数（忽略末尾的空行）
            actual_max_row = 0
            for row in range(1, ws_in.max_row + 1):
                row_has_data = False
                for col in range(1, ws_in.max_column + 1):
                    if ws_in.cell(row=row, column=col).value is not None:
                        row_has_data = True
                        break
                if row_has_data:
                    actual_max_row = row
            
            # 如果没有实际数据，跳过当前文件
            if actual_max_row == 0:
                continue
            
            # 确定起始行：
            # - 如果是第一个文件，从第1行开始
            # - 如果是后续文件且设置了表头行数，从表头行数+1行开始
            # - 否则从第1行开始
            start_row = 1
            if i > 0 and header_rows > 0:
                start_row = header_rows + 1
            
            # 复制数据和格式（只复制实际有数据的行）
            for row in ws_in.iter_rows(min_row=start_row, max_row=actual_max_row, min_col=1, max_col=ws_in.max_column):
                for cell in row:
                    # 计算目标单元格位置
                    # 注意：当跳过表头时，需要调整行号计算
                    if i > 0 and header_rows > 0:
                        target_row = cell.row + row_offset - header_rows
                    else:
                        target_row = cell.row + row_offset
                    target_col = cell.column
                    target_cell = ws_out.cell(row=target_row, column=target_col, value=cell.value)
                    
                    # 复制单元格样式
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.alignment = copy(cell.alignment)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = cell.number_format
                        target_cell.protection = copy(cell.protection)
            
            # 应用列宽
            for col, width in formats['column_widths'].items():
                ws_out.column_dimensions[col].width = width
            
            # 应用行高（只应用实际复制的行）
            for row, height in formats['row_heights'].items():
                if row >= start_row and row <= actual_max_row:
                    # 调整目标行号
                    if i > 0 and header_rows > 0:
                        target_row = row + row_offset - header_rows
                    else:
                        target_row = row + row_offset
                    ws_out.row_dimensions[target_row].height = height
            
            # 应用合并单元格（只应用实际复制的行范围内的合并）
            for merge_range in formats['merged_cells']:
                from openpyxl.utils import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(merge_range)
                
                # 检查合并单元格是否在实际有数据的行范围内
                if max_row <= actual_max_row:
                    # 如果是后续文件且设置了表头行数，且合并单元格在表头范围内，则跳过
                    if i > 0 and header_rows > 0 and min_row <= header_rows:
                        continue
                    
                    # 调整合并单元格的行号
                    if i > 0 and header_rows > 0:
                        adjusted_merge = adjust_merge_range(merge_range, row_offset - header_rows, 0)
                    else:
                        adjusted_merge = adjust_merge_range(merge_range, row_offset, 0)
                    ws_out.merge_cells(adjusted_merge)
            
            # 更新行偏移量（只增加实际复制的行数）
            if i > 0 and header_rows > 0:
                row_offset += (actual_max_row - start_row + 1)
            else:
                row_offset += (actual_max_row - start_row + 1)
    
    else:
        # 水平合并（追加列）
        col_offset = 0
        for file_path in file_list:
            wb_in = load_workbook(file_path)
            ws_in = wb_in.active
            
            # 读取源工作表的格式信息
            formats = read_worksheet_formats(ws_in)
            
            # 复制数据和格式
            for row in ws_in.iter_rows(min_row=1, max_row=ws_in.max_row, min_col=1, max_col=ws_in.max_column):
                for cell in row:
                    # 计算目标单元格位置
                    target_row = cell.row
                    target_col = cell.column + col_offset
                    target_cell = ws_out.cell(row=target_row, column=target_col, value=cell.value)
                    
                    # 复制单元格样式
                    if cell.has_style:
                        target_cell.font = copy(cell.font)
                        target_cell.alignment = copy(cell.alignment)
                        target_cell.border = copy(cell.border)
                        target_cell.fill = copy(cell.fill)
                        target_cell.number_format = cell.number_format
                        target_cell.protection = copy(cell.protection)
            
            # 应用列宽
            for col, width in formats['column_widths'].items():
                # 调整列号
                col_idx = ws_in[col][0].column + col_offset
                col_letter = get_column_letter(col_idx)
                ws_out.column_dimensions[col_letter].width = width
            
            # 应用行高
            for row, height in formats['row_heights'].items():
                ws_out.row_dimensions[row].height = height
            
            # 应用合并单元格
            for merge_range in formats['merged_cells']:
                # 调整合并单元格的列号
                adjusted_merge = adjust_merge_range(merge_range, 0, col_offset)
                ws_out.merge_cells(adjusted_merge)
            
            # 更新列偏移量
            col_offset += ws_in.max_column
    
    # 保存结果
    wb_out.save(output_path)
    return True, f"合并完成！输出文件：{output_path}"


def read_worksheet_formats(ws):
    """
    读取工作表的格式信息
    
    Args:
        ws: openpyxl工作表对象
        
    Returns:
        dict: 包含格式信息的字典，包括：
            - merged_cells: 合并单元格范围列表
            - column_widths: 列宽字典
            - row_heights: 行高字典
    """
    formats = {
        'merged_cells': [],
        'column_widths': {},
        'row_heights': {}
    }
    
    # 读取合并单元格
    for merge_cell in ws.merged_cells.ranges:
        formats['merged_cells'].append(str(merge_cell))
    
    # 读取列宽
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if ws.column_dimensions[col_letter].width is not None:
            formats['column_widths'][col_letter] = ws.column_dimensions[col_letter].width
    
    # 读取行高
    for row in range(1, ws.max_row + 1):
        if ws.row_dimensions[row].height is not None:
            formats['row_heights'][row] = ws.row_dimensions[row].height
    
    return formats


def adjust_merge_range(merge_range, row_offset, col_offset):
    """
    调整合并单元格范围的行号和列号
    
    Args:
        merge_range (str): 合并单元格范围字符串，如 "A1:B2"
        row_offset (int): 行偏移量
        col_offset (int): 列偏移量
        
    Returns:
        str: 调整后的合并单元格范围字符串
    """
    from openpyxl.utils import range_boundaries
    
    # 解析范围边界
    min_col, min_row, max_col, max_row = range_boundaries(merge_range)
    
    # 应用偏移量
    min_row += row_offset
    max_row += row_offset
    min_col += col_offset
    max_col += col_offset
    
    # 转换回范围字符串
    min_col_letter = get_column_letter(min_col)
    max_col_letter = get_column_letter(max_col)
    
    return f"{min_col_letter}{min_row}:{max_col_letter}{max_row}"


class TableMergeFunction(BaseFileProcessorFunction):
    """表格合并功能"""
    
    def __init__(self, parent=None):
        description = (
            "📢 <b>功能说明：</b><br>" 
            "合并多个表格文件，支持拖入文件和添加目录<br>" 
            "可以调整合并顺序，设置默认路径和合并方式"
        )
        super().__init__("表格合并", description, parent)
        self.init_ui()
    
    def init_ui(self):
        """初始化UI界面"""
        # 直接使用父类的contentLayout
        self.contentLayout.setContentsMargins(0, 0, 0, 0)
        self.contentLayout.setSpacing(15)
        
        # 创建文件列表区域
        self.create_file_list_section()
        
        # 创建合并设置区域
        self.create_merge_settings_section()
        
        # 创建输出设置区域
        self.create_output_settings_section()
    
    def create_file_list_section(self):
        """创建文件列表区域"""
        # 创建文件列表组
        self.file_list_group = QGroupBox("文件列表")
        self.file_list_layout = QVBoxLayout(self.file_list_group)
        self.file_list_layout.setSpacing(10)
        
        # 创建列表描述
        self.list_desc_label = QLabel("可以将表格文件拖入列表，或通过按钮添加文件/目录")
        self.list_desc_label.setStyleSheet("font-size: 12px; color: #666;")
        self.file_list_layout.addWidget(self.list_desc_label)
        
        # 创建文件列表表格
        self.file_list_widget = TableWidget()
        self.file_list_widget.setColumnCount(1)
        self.file_list_widget.setHorizontalHeaderLabels(["文件路径"])
        from PyQt6.QtWidgets import QHeaderView
        self.file_list_widget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.file_list_widget.setAlternatingRowColors(True)
        self.file_list_widget.setFixedHeight(200)
        self.file_list_widget.setAcceptDrops(True)
        self.file_list_widget.dragEnterEvent = self.on_drag_enter
        self.file_list_widget.dropEvent = self.on_drop
        self.file_list_layout.addWidget(self.file_list_widget)
        
        # 创建列表操作按钮区域
        self.list_buttons_layout = QHBoxLayout()
        self.list_buttons_layout.setSpacing(10)
        
        # 添加文件按钮
        self.add_file_btn = PushButton("添加文件", icon=FluentIcon.ADD_TO)
        self.add_file_btn.clicked.connect(self.add_file)
        
        # 添加目录按钮
        self.add_dir_btn = PushButton("添加目录", icon=FluentIcon.FOLDER)
        self.add_dir_btn.clicked.connect(self.add_directory)
        
        # 移除选中按钮
        self.remove_selected_btn = PushButton("移除选中", icon=FluentIcon.DELETE)
        self.remove_selected_btn.clicked.connect(self.remove_selected)
        
        # 上移按钮
        self.move_up_btn = PushButton("上移", icon=FluentIcon.UP)
        self.move_up_btn.clicked.connect(self.move_up)
        
        # 下移按钮
        self.move_down_btn = PushButton("下移", icon=FluentIcon.DOWN)
        self.move_down_btn.clicked.connect(self.move_down)
        
        # 清空列表按钮
        self.clear_list_btn = PushButton("清空列表", icon=FluentIcon.DELETE)
        self.clear_list_btn.clicked.connect(self.clear_list)
        
        # 添加按钮到布局
        self.list_buttons_layout.addWidget(self.add_file_btn)
        self.list_buttons_layout.addWidget(self.add_dir_btn)
        self.list_buttons_layout.addWidget(self.remove_selected_btn)
        self.list_buttons_layout.addWidget(self.move_up_btn)
        self.list_buttons_layout.addWidget(self.move_down_btn)
        self.list_buttons_layout.addWidget(self.clear_list_btn)
        self.file_list_layout.addLayout(self.list_buttons_layout)
        
        # 添加到contentLayout
        self.contentLayout.addWidget(self.file_list_group)
    
    def create_merge_settings_section(self):
        """创建合并设置区域"""
        # 创建合并设置组
        self.merge_settings_group = QGroupBox("合并设置")
        self.merge_settings_layout = QVBoxLayout(self.merge_settings_group)
        self.merge_settings_layout.setSpacing(10)
        
        # 合并方式设置
        self.merge_type_layout = QHBoxLayout()
        self.merge_type_label = QLabel("合并方式：")
        self.merge_type_combo = ComboBox()
        self.merge_type_combo.addItems(["垂直合并（追加行）", "水平合并（追加列）"])
        self.merge_type_combo.setCurrentIndex(0)
        # 添加信号连接，监听合并方式变化
        self.merge_type_combo.currentIndexChanged.connect(self.on_merge_type_changed)
        
        self.merge_type_layout.addWidget(self.merge_type_label)
        self.merge_type_layout.addWidget(self.merge_type_combo)
        self.merge_type_layout.addStretch(1)
        self.merge_settings_layout.addLayout(self.merge_type_layout)
        
        # 表头处理设置 - 改为数值输入
        self.header_layout = QHBoxLayout()
        self.header_label = QLabel("表头行数：")
        self.header_rows_edit = LineEdit()
        self.header_rows_edit.setText("0")
        self.header_rows_edit.setPlaceholderText("输入表头行数，0表示保留所有表头")
        self.header_rows_edit.setMaximumWidth(100)
        
        self.header_layout.addWidget(self.header_label)
        self.header_layout.addWidget(self.header_rows_edit)
        self.header_layout.addWidget(QLabel("行"))
        self.header_layout.addStretch(1)
        self.merge_settings_layout.addLayout(self.header_layout)
        
        # 初始化表头设置的可见性
        self.on_merge_type_changed(0)
        
        # 添加到contentLayout
        self.contentLayout.addWidget(self.merge_settings_group)
    
    def create_output_settings_section(self):
        """创建输出设置区域"""
        # 创建输出设置组
        self.output_group = QGroupBox("输出设置")
        self.output_layout = QVBoxLayout(self.output_group)
        self.output_layout.setSpacing(10)
        
        # 输出路径设置
        self.output_path_layout = QHBoxLayout()
        self.output_path_label = QLabel("输出路径：")
        self.output_path_edit = LineEdit()
        self.output_path_edit.setPlaceholderText("选择输出文件路径")
        
        self.browse_output_btn = PushButton("浏览", icon=FluentIcon.FOLDER)
        self.browse_output_btn.clicked.connect(self.browse_output_path)
        
        self.output_path_layout.addWidget(self.output_path_label)
        self.output_path_layout.addWidget(self.output_path_edit, 1)
        self.output_path_layout.addWidget(self.browse_output_btn)
        self.output_layout.addLayout(self.output_path_layout)
        
        # 输出文件名设置
        self.output_filename_layout = QHBoxLayout()
        self.output_filename_label = QLabel("输出文件名：")
        self.output_filename_edit = LineEdit()
        
        self.output_filename_layout.addWidget(self.output_filename_label)
        self.output_filename_layout.addWidget(self.output_filename_edit, 1)
        self.output_layout.addLayout(self.output_filename_layout)
        
        # 添加执行按钮
        self.execute_btn = PrimaryPushButton("开始合并", icon=FluentIcon.SEND)
        self.execute_btn.clicked.connect(self.execute_merge)
        self.output_layout.addWidget(self.execute_btn, alignment=Qt.AlignmentFlag.AlignCenter)
        
        # 添加到contentLayout
        self.contentLayout.addWidget(self.output_group)
    
    def on_drag_enter(self, event: QDragEnterEvent):
        """处理拖入事件"""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
    
    def on_drop(self, event: QDropEvent):
        """处理拖放事件"""
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if os.path.isfile(file_path):
                if file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xls') or file_path.lower().endswith('.csv'):
                    self.add_file_to_list(file_path)
    
    def add_file(self):
        """添加文件到列表"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "选择表格文件", self.get_default_path(), 
            "表格文件 (*.xlsx *.xls *.csv)"
        )
        for file in files:
            self.add_file_to_list(file)
    
    def add_directory(self):
        """添加目录中的表格文件到列表"""
        dir_path = QFileDialog.getExistingDirectory(
            self, "选择目录", self.get_default_path()
        )
        if dir_path:
            for file in os.listdir(dir_path):
                file_path = os.path.join(dir_path, file)
                if os.path.isfile(file_path):
                    if file_path.lower().endswith('.xlsx') or file_path.lower().endswith('.xls') or file_path.lower().endswith('.csv'):
                        self.add_file_to_list(file_path)
    
    def add_file_to_list(self, file_path):
        """将文件添加到列表"""
        # 检查文件是否已在列表中
        for row in range(self.file_list_widget.rowCount()):
            if self.file_list_widget.item(row, 0).text() == file_path:
                return
        
        # 添加到表格
        row = self.file_list_widget.rowCount()
        self.file_list_widget.insertRow(row)
        self.file_list_widget.setItem(row, 0, QTableWidgetItem(file_path))
        
        # 更新输出路径和输出文件名
        if row == 0:  # 只有当添加第一个文件时更新
            # 设置输出路径为默认路径
            self.output_path_edit.setText(self.get_default_path())
            # 设置默认输出文件名
            self.output_filename_edit.setText("merged_table.xlsx")
    
    def remove_selected(self):
        """移除选中的项目"""
        selected_rows = []
        for item in self.file_list_widget.selectedItems():
            selected_rows.append(item.row())
        
        # 按降序删除，避免索引问题
        for row in sorted(selected_rows, reverse=True):
            self.file_list_widget.removeRow(row)
        
        # 检查列表是否为空
        if self.file_list_widget.rowCount() == 0:
            # 清空输出路径和输出文件名
            self.output_path_edit.setText("")
            self.output_filename_edit.setText("")
    
    def move_up(self):
        """将选中的项目上移"""
        current_item = self.file_list_widget.currentItem()
        if not current_item:
            return
        current_row = current_item.row()
        if current_row > 0:
            # 获取当前行数据
            file_path = self.file_list_widget.item(current_row, 0).text()
            
            # 移除当前行
            self.file_list_widget.removeRow(current_row)
            
            # 插入到上一行
            self.file_list_widget.insertRow(current_row - 1)
            self.file_list_widget.setItem(current_row - 1, 0, QTableWidgetItem(file_path))
            
            # 设置当前行
            self.file_list_widget.setCurrentItem(self.file_list_widget.item(current_row - 1, 0))
    
    def move_down(self):
        """将选中的项目下移"""
        current_item = self.file_list_widget.currentItem()
        if not current_item:
            return
        current_row = current_item.row()
        if current_row < self.file_list_widget.rowCount() - 1:
            # 获取当前行数据
            file_path = self.file_list_widget.item(current_row, 0).text()
            
            # 移除当前行
            self.file_list_widget.removeRow(current_row)
            
            # 插入到下一行
            self.file_list_widget.insertRow(current_row + 1)
            self.file_list_widget.setItem(current_row + 1, 0, QTableWidgetItem(file_path))
            
            # 设置当前行
            self.file_list_widget.setCurrentItem(self.file_list_widget.item(current_row + 1, 0))
    
    def clear_list(self):
        """清空列表"""
        self.file_list_widget.setRowCount(0)
        # 清空输出路径和输出文件名
        self.output_path_edit.setText("")
        self.output_filename_edit.setText("")
    
    def get_default_path(self):
        """获取默认路径"""
        if self.file_list_widget.rowCount() > 0:
            first_file = self.file_list_widget.item(0, 0).text()
            return os.path.dirname(first_file)
        return os.getcwd()
    
    def browse_output_path(self):
        """浏览输出路径"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存合并后的表格", self.get_default_path(), 
            "Excel文件 (*.xlsx);;CSV文件 (*.csv)"
        )
        if file_path:
            self.output_path_edit.setText(file_path)
            # 自动提取文件名
            self.output_filename_edit.setText(os.path.basename(file_path))
    
    def execute_merge(self):
        """执行表格合并"""
        # 获取文件列表
        file_list = []
        for i in range(self.file_list_widget.rowCount()):
            file_list.append(self.file_list_widget.item(i, 0).text())
        
        if not file_list:
            QMessageBox.warning(self, "警告", "请先添加要合并的表格文件！")
            return
        
        # 获取合并设置
        merge_type = 'vertical' if self.merge_type_combo.currentIndex() == 0 else 'horizontal'
        
        # 获取表头行数设置
        try:
            header_rows = int(self.header_rows_edit.text().strip())
            if header_rows < 0:
                header_rows = 0
        except ValueError:
            header_rows = 0
        
        # 获取输出路径
        output_path = self.output_path_edit.text().strip()
        output_filename = self.output_filename_edit.text().strip()
        
        # 验证输出路径
        if not output_path or not output_filename:
            # 使用默认输出路径和文件名
            default_dir = self.get_default_path()
            output_path = os.path.join(default_dir, output_filename if output_filename else "merged_table.xlsx")
        elif os.path.isdir(output_path):
            # 如果输出路径是目录，自动添加文件名
            output_path = os.path.join(output_path, output_filename if output_filename else "merged_table.xlsx")
        
        try:
            self.showProgress("正在合并表格...")
            
            # 执行合并
            success, message = merge_tables(file_list, output_path, merge_type, header_rows=header_rows)
            
            if success:
                self.showSuccess(message)
            else:
                self.showError(message)
        except Exception as e:
            self.showError(f"合并过程中发生错误：{str(e)}")
    
    def on_merge_type_changed(self, index):
        """
        合并方式变化时的处理函数
        
        Args:
            index (int): 合并方式索引，0表示垂直合并，1表示水平合并
        """
        # 垂直合并时显示表头行数设置，水平合并时隐藏
        is_vertical = (index == 0)
        
        # 遍历header_layout中的所有小部件，设置可见性
        for i in range(self.header_layout.count()):
            widget = self.header_layout.itemAt(i).widget()
            if widget:
                widget.setVisible(is_vertical)
    
    def execute(self):
        """执行功能（BaseFunction要求实现的方法）"""
        self.execute_merge()
